from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QWidget, QPushButton, QVBoxLayout, QLabel, QLineEdit, QTextEdit
from PyQt6.QtGui import QFont
import os
import sqlite3
from openpyxl import Workbook
from functions import delete_end_of_string, get_activity_list_path, get_activity_description_path, get_all_classes_path, get_all_activities_path, get_db_path, get_new_students_path, set_to_str
from Button import Button
from TableStudents import TableStudents
from InputWindow import InputWindow
from ErrorWindow import ErrorWindow


class Window(QWidget):
    def __init__(self):
        super().__init__()
        self.input_win = None
        self.error_win = None
        self.table_students = TableStudents()
        self.create_widgets()
        self.open_main_page()
    
    def create_widgets(self):
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)
        self.create_main_page_buttons()
        self.crate_grade_selection_buttons()
        self.create_class_selection_buttons()
        self.create_activity_selection_buttons()
        self.create_activity_inner_pages()
        self.create_activity_description_pages()
        self.hide_all_buttons()

    def create_main_page_buttons(self):
        self.main_page_buttons = []

        button_student_lists = Button("Список классов", self)
        button_student_lists.clicked.connect(self.open_grade_selection_page)
        self.main_page_buttons.append(button_student_lists)

        button_activity_lists = Button("Мероприятия", self)
        button_activity_lists.clicked.connect(self.open_activity_selection_page)
        self.main_page_buttons.append(button_activity_lists)
        
        button_export = Button("Экспортировать таблицу", self)
        button_export.clicked.connect(self.export_table)
        self.main_page_buttons.append(button_export)

        button_add_student = Button("Добавить ученика", self)
        button_add_student.clicked.connect(self.open_add_student_page)
        self.main_page_buttons.append(button_add_student)

        for button in self.main_page_buttons:
            self.layout.addWidget(button, alignment=Qt.AlignmentFlag.AlignCenter)

    def crate_grade_selection_buttons(self):
        self.grade_selection_buttons = []
        button_all = Button("Все ученики", self)
        button_all.clicked.connect(lambda state, x = 0 : self.open_class_selection_page(x))
        self.grade_selection_buttons.append(button_all)

        for grade in range(5, 12):
            button = Button(str(grade), self)
            button.clicked.connect(lambda state, x = grade : self.open_class_selection_page(x))
            self.grade_selection_buttons.append(button)
        
        button_go_back_from_grade_selection = Button("Назад", self)
        button_go_back_from_grade_selection.clicked.connect(self.open_main_page)
        self.grade_selection_buttons.append(button_go_back_from_grade_selection)

        for button in self.grade_selection_buttons:
            self.layout.addWidget(button, alignment=Qt.AlignmentFlag.AlignCenter)

    def create_class_selection_buttons(self):
        self.class_selection_buttons = [[] for i in range(12)]
        with open(get_all_classes_path(), "r", encoding="utf-8") as all_classes:
            for cur_class_name in all_classes:
                cur_class_name = delete_end_of_string(cur_class_name)
                grade = int(cur_class_name[0:1])
                if len(cur_class_name) == 4:
                    grade = int(cur_class_name[0:2])
                button = Button(cur_class_name, self)
                button.clicked.connect(lambda state, x = cur_class_name : self.open_table_class(x))
                self.class_selection_buttons[grade].append(button)

        for grade in range(12):
            button_go_back_from_class_selection = Button("Назад", self)
            button_go_back_from_class_selection.clicked.connect(self.open_grade_selection_page)
            self.class_selection_buttons[grade].append(button_go_back_from_class_selection)

            for button in self.class_selection_buttons[grade]:
                self.layout.addWidget(button, alignment=Qt.AlignmentFlag.AlignCenter)

    def create_activity_selection_buttons(self):
        self.activity_selection_buttons = []

        button_create_activity = Button("Добавить мероприятие", self)
        button_create_activity.clicked.connect(self.open_create_new_activity_window)
        self.activity_selection_buttons.append(button_create_activity)

        with open(get_all_activities_path(), "r", encoding="utf-8") as all_activities:
            for cur_activity_name in all_activities:
                cur_activity_name = delete_end_of_string(cur_activity_name)
                button = Button(cur_activity_name, self)
                button.clicked.connect(lambda state, x = cur_activity_name : self.open_activity_inner_page(x))
                self.activity_selection_buttons.append(button)

        button_go_back_from_activity_selection = Button("Назад", self)
        button_go_back_from_activity_selection.clicked.connect(self.open_main_page)
        self.activity_selection_buttons.append(button_go_back_from_activity_selection)

        for button in self.activity_selection_buttons:
            self.layout.addWidget(button, alignment=Qt.AlignmentFlag.AlignCenter)

    def create_activity_inner_pages(self):
        self.activity_inner_page_widgets = {}
        with open(get_all_activities_path(), "r", encoding="utf-8") as all_activities:
            for cur_activity_name in all_activities:
                cur_activity_name = delete_end_of_string(cur_activity_name)
                self.create_activity_inner_page(cur_activity_name)

    def create_activity_inner_page(self, activity_name):
        self.activity_inner_page_widgets[activity_name] = []

        label_activity_name = QLabel(activity_name, self)
        label_activity_name.setFont(QFont("Helvetica [Cronyx]", 18))
        self.activity_inner_page_widgets[activity_name].append(label_activity_name)

        button_edit = Button("Редактировать", self)
        button_edit.clicked.connect(lambda state, x = activity_name : self.open_activity_edit_window(x))
        self.activity_inner_page_widgets[activity_name].append(button_edit)

        button_participants_list = Button("Участники", self)
        button_participants_list.clicked.connect(lambda state, x = activity_name : self.open_table_activity(x))
        self.activity_inner_page_widgets[activity_name].append(button_participants_list)

        button_description = Button("Описание", self)
        button_description.clicked.connect(lambda state, x = activity_name : self.open_activity_description(x))
        self.activity_inner_page_widgets[activity_name].append(button_description)

        button_delete = Button("Удалить", self)
        button_delete.clicked.connect(lambda state, x = activity_name : self.open_delete_activity_window(x))
        self.activity_inner_page_widgets[activity_name].append(button_delete)

        button_go_back_from_activity_inner_page = Button("Назад", self)
        button_go_back_from_activity_inner_page.clicked.connect(self.open_activity_selection_page)
        self.activity_inner_page_widgets[activity_name].append(button_go_back_from_activity_inner_page)

        for widget in self.activity_inner_page_widgets[activity_name]:
            self.layout.addWidget(widget, alignment=Qt.AlignmentFlag.AlignCenter)

    def create_activity_description_pages(self):
        self.activity_description_page_widgets = {}
        with open(get_all_activities_path(), "r", encoding="utf-8") as all_activities:
            for cur_activity_name in all_activities:
                cur_activity_name = delete_end_of_string(cur_activity_name)
                with open(get_activity_description_path(cur_activity_name), "r", encoding="utf-8") as description:
                    self.create_activity_description_page(cur_activity_name, description.read())

    def create_activity_description_page(self, activity_name, activity_description):
        self.activity_description_page_widgets[activity_name] = []

        label_description = QLabel(activity_description, self)
        label_description.setFont(QFont("Helvetica [Cronyx]", 12))
        self.layout.addWidget(label_description, alignment=Qt.AlignmentFlag.AlignCenter)
        self.activity_description_page_widgets[activity_name].append(label_description)

        button_go_back_from_activity_description = Button("Назад", self)
        button_go_back_from_activity_description.clicked.connect(lambda state, x = activity_name : self.open_activity_inner_page(x))
        self.layout.addWidget(button_go_back_from_activity_description, alignment=Qt.AlignmentFlag.AlignCenter)
        self.activity_description_page_widgets[activity_name].append(button_go_back_from_activity_description)

    def hide_all_buttons(self):
        self.table_students.hide()
        if self.error_win:
            self.error_win.close()
        if self.input_win:
            self.input_win.close()
        for button in self.main_page_buttons:
            button.hide()
        for button in self.grade_selection_buttons:
            button.hide()
        for grade in range(12):
            for button in self.class_selection_buttons[grade]:
                button.hide()
        for button in self.activity_selection_buttons:
            button.hide()
        for widgets in self.activity_inner_page_widgets.values():
            for widget in widgets:
                widget.hide()
        for widgets in self.activity_description_page_widgets.values():
            for widget in widgets:
                widget.hide()

    def open_main_page(self):
        self.hide_all_buttons()
        for button in self.main_page_buttons:
            button.show()

    def open_grade_selection_page(self):
        self.hide_all_buttons()
        for button in self.grade_selection_buttons:
            button.show()

    def open_class_selection_page(self, grade):
        if (grade == 0):
            self.open_table_all_students()
            return
        self.hide_all_buttons()
        for button in self.class_selection_buttons[grade]:
            button.show()

    def open_activity_selection_page(self):
        self.hide_all_buttons()
        for button in self.activity_selection_buttons:
            button.show()

    def open_activity_inner_page(self, activity_name):
        self.hide_all_buttons()
        for widget in self.activity_inner_page_widgets[activity_name]:
            widget.show()

    def open_activity_description(self, activity_name):
        self.hide_all_buttons()
        for widget in self.activity_description_page_widgets[activity_name]:
            widget.show()

    def open_table_all_students(self):
        self.table_students.hide()
        self.table_students.setWindowTitle("Все ученики")
        self.table_students.show()

        for i in range(self.table_students.table_model.columnCount()):
            self.table_students.showColumn(i)
        for i in range(self.table_students.table_model.rowCount()):
            self.table_students.showRow(i)
            self.table_students.set_value_in_cell(i, 3, len(self.table_students.student_activities[i]))
            self.table_students.set_value_in_cell(i, 4, set_to_str(self.table_students.student_activities[i]))

        self.table_students.hideColumn(0)

    def open_table_class(self, class_name):
        self.table_students.hide()
        self.table_students.setWindowTitle(f"Список учеников класса {class_name}")
        self.table_students.show()

        for i in range(self.table_students.table_model.columnCount()):
            self.table_students.showColumn(i)
        for i in range(self.table_students.table_model.rowCount()):
            self.table_students.showRow(i)
            index = self.table_students.table_model.index(i, 2)
            cur_class_name = self.table_students.table_model.data(index)
            if (cur_class_name != class_name):
                self.table_students.hideRow(i)
            else:
                self.table_students.set_value_in_cell(i, 3, len(self.table_students.student_activities[i]))
                self.table_students.set_value_in_cell(i, 4, set_to_str(self.table_students.student_activities[i]))

        self.table_students.hideColumn(0)
        self.table_students.hideColumn(2)

    def open_table_activity(self, activity_name):
        self.table_students.hide()
        self.table_students.setWindowTitle(f"Список участников мероприятия '{activity_name}'")
        self.table_students.show()

        for i in range(self.table_students.table_model.columnCount()):
            self.table_students.showColumn(i)
        for i in range(self.table_students.table_model.rowCount()):
            self.table_students.showRow(i)
            if (not (activity_name in self.table_students.student_activities[i])):
                self.table_students.hideRow(i)
            else:
                self.table_students.set_value_in_cell(i, 3, len(self.table_students.student_activities[i]))
                self.table_students.set_value_in_cell(i, 4, set_to_str(self.table_students.student_activities[i]))

        self.table_students.hideColumn(0)

    def open_create_new_activity_window(self):
        self.hide()

        self.input_win = InputWindow(self)
        self.input_win.setGeometry(400, 250, 450, 600)

        label_name = QLabel("Введите название мероприятия:", self)
        label_name.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label_name, alignment=Qt.AlignmentFlag.AlignCenter)

        self.input_win.input_name = QLineEdit()
        self.input_win.layout.addWidget(self.input_win.input_name, alignment=Qt.AlignmentFlag.AlignCenter)

        label_description = QLabel("Введите описание мероприятия (не обязательно):", self)
        label_description.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label_description, alignment=Qt.AlignmentFlag.AlignCenter)

        self.input_win.input_description = QTextEdit()
        self.input_win.input_description.setMinimumWidth(380)
        self.input_win.layout.addWidget(self.input_win.input_description, alignment=Qt.AlignmentFlag.AlignCenter)

        label_participants = QLabel("Введите ФИО и класс участников мероприятия:", self)
        label_participants.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label_participants, alignment=Qt.AlignmentFlag.AlignCenter)

        self.input_win.input_participants = QTextEdit()
        self.input_win.input_participants.setMinimumWidth(380)
        self.input_win.layout.addWidget(self.input_win.input_participants, alignment=Qt.AlignmentFlag.AlignCenter)

        button_go_back = QPushButton("Отмена")
        button_go_back.clicked.connect(self.return_from_activity_creation)
        self.input_win.layout.addWidget(button_go_back)

        button_commit = QPushButton("Подтвердить")
        button_commit.clicked.connect(self.create_new_activity)
        self.input_win.layout.addWidget(button_commit)

        self.input_win.show()

    def check_new_activity(self, activity_name, activity_participants):
        if len(activity_name) == 0:
            self.error_win = ErrorWindow("Пожалуйста, введите название мероприятия.")
            return False

        with open(get_all_activities_path(), "r", encoding="utf-8") as all_activities:
            for cur_activity_name in all_activities:
                cur_activity_name = delete_end_of_string(cur_activity_name)
                if (activity_name == cur_activity_name):
                    self.error_win = ErrorWindow(f"Мероприятие с таким названием уже существует.\nПожалуйста, выберите другое название.")
                    return False
                
        str_ind = 0
        student_line = {}
        for student_name in activity_participants.split('\n'):
            str_ind += 1
            if (len(student_name) == 0):
                continue

            if (student_line.get(student_name) is None):
                student_line[student_name] = str_ind
            else:
                self.error_win = ErrorWindow(f"Ученик '{student_name}' встречается\nв двух строках, а именно в {student_line[student_name]}-ой и {str_ind}-ой.\nПожалуйста, проверьте, правильно ли Вы ввели данные.")
                return False
            
            if (self.table_students.student_id.get(student_name) is None):
                self.error_win = ErrorWindow(f"Ученика '{student_name}'\n(строка {str_ind}) нет в базе.\nУбедитесь, что данные введены корректно.\nФормат ввода:\n'Фамилия имя отчество класс' (без кавычек)")
                return False

        return True

    def create_new_activity(self):
        self.error_win = None

        activity_name = self.input_win.input_name.text()
        activity_description = self.input_win.input_description.toPlainText()
        activity_participants = self.input_win.input_participants.toPlainText()

        if (self.check_new_activity(activity_name, activity_participants)):
            self.activity_selection_buttons[-1].setText(activity_name)
            self.activity_selection_buttons[-1].clicked.connect(lambda state, x = activity_name : self.open_activity_inner_page(x))

            button_go_back_from_activity_selection = Button("Назад", self)
            button_go_back_from_activity_selection.clicked.connect(self.open_main_page)
            self.layout.addWidget(button_go_back_from_activity_selection, alignment=Qt.AlignmentFlag.AlignCenter)
            self.activity_selection_buttons.append(button_go_back_from_activity_selection)
            self.hide_all_buttons()

            self.create_activity_inner_page(activity_name)
            self.create_activity_description_page(activity_name, activity_description)

            with open(get_all_activities_path(), "a", encoding="utf-8") as all_activities:
                all_activities.write(activity_name + '\n')

            with open(get_activity_list_path(activity_name), "w", encoding="utf-8") as participants_list:
                for student_name in activity_participants.split('\n'):
                    if (len(student_name) == 0):
                        continue
                    participants_list.write(student_name + '\n')
                    self.table_students.student_activities[self.table_students.student_id[student_name]].add(activity_name)
                
            with open(get_activity_description_path(activity_name), "w", encoding="utf-8") as description:
                description.write(activity_description)

            self.input_win.close()
            self.open_activity_selection_page()

    def return_from_activity_creation(self):
        if self.error_win:
            self.error_win.close()
        self.input_win.close()
        self.show()
        self.open_activity_selection_page()

    def open_activity_edit_window(self, activity_name):
        self.hide()

        self.input_win = InputWindow(self)
        self.input_win.setGeometry(400, 150, 450, 800)

        label_description = QLabel("Введите описание мероприятия:\n(Если не хотите менять описание мероприятия,\nоставьте поле пустым.)", self)
        label_description.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label_description, alignment=Qt.AlignmentFlag.AlignCenter)

        self.input_win.input_description = QTextEdit()
        self.input_win.input_description.setMinimumWidth(380)
        self.input_win.layout.addWidget(self.input_win.input_description, alignment=Qt.AlignmentFlag.AlignCenter)

        label_add_participants = QLabel("Введите ФИО и класс участников мероприятия,\nкоторых Вы хотите добавить:", self)
        label_add_participants.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label_add_participants, alignment=Qt.AlignmentFlag.AlignCenter)

        self.input_win.input_add_participants = QTextEdit()
        self.input_win.input_add_participants.setMinimumWidth(380)
        self.input_win.layout.addWidget(self.input_win.input_add_participants, alignment=Qt.AlignmentFlag.AlignCenter)

        label_del_participants = QLabel("Введите ФИО и класс участников мероприятия,\nкоторых Вы хотите удалить:", self)
        label_del_participants.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label_del_participants, alignment=Qt.AlignmentFlag.AlignCenter)

        self.input_win.input_del_participants = QTextEdit()
        self.input_win.input_del_participants.setMinimumWidth(380)
        self.input_win.layout.addWidget(self.input_win.input_del_participants, alignment=Qt.AlignmentFlag.AlignCenter)

        button_go_back = QPushButton("Отмена")
        button_go_back.clicked.connect(lambda state, x = activity_name : self.return_from_activity_edit_or_deletion(x))
        self.input_win.layout.addWidget(button_go_back)

        button_commit = QPushButton("Подтвердить")
        button_commit.clicked.connect(lambda state, x = activity_name : self.edit_activity(x))
        self.input_win.layout.addWidget(button_commit)

        self.input_win.show()

    def check_edited_activity(self, activity_participants):
        str_ind = 0
        student_line = {}
        for student_name in activity_participants.split('\n'):
            str_ind += 1
            if (len(student_name) == 0):
                continue

            if (student_line.get(student_name) is None):
                student_line[student_name] = str_ind
            else:
                self.error_win = ErrorWindow(f"Ученик '{student_name}' встречается\nв двух строках, а именно в {student_line[student_name]}-ой и {str_ind}-ой.\nПожалуйста, проверьте, правильно ли Вы ввели данные.")
                return False
            
            if (self.table_students.student_id.get(student_name) is None):
                self.error_win = ErrorWindow(f"Ученика '{student_name}' (строка {str_ind}) нет в базе.\nУбедитесь, что данные введены корректно.\nФормат ввода:\n'Фамилия имя отчество класс' (без кавычек)")
                return False
        
        return True

    def edit_activity(self, activity_name):
        self.error_win = None

        activity_description = self.input_win.input_description.toPlainText()
        activity_add_participants = self.input_win.input_add_participants.toPlainText()
        activity_del_participants = self.input_win.input_del_participants.toPlainText()

        if (self.check_edited_activity(activity_add_participants) and self.check_edited_activity(activity_del_participants)):    
            activity_add_participants = activity_add_participants.split('\n')
            activity_del_participants = activity_del_participants.split('\n')

            new_participants = []
            with open(get_activity_list_path(activity_name), "r", encoding="utf-8") as participants_list:
                for line in participants_list:
                    line = delete_end_of_string(line)
                    if not (line in activity_del_participants):
                        new_participants.append(line + '\n')

                for student_name in activity_add_participants:
                    if (len(student_name) == 0 
                        or activity_name in self.table_students.student_activities[self.table_students.student_id[student_name]]
                        or student_name in activity_del_participants):
                        continue
                    new_participants.append(student_name + '\n')
                    self.table_students.student_activities[self.table_students.student_id[student_name]].add(activity_name)

                for student_name in activity_del_participants:
                    if (len(student_name) == 0):
                        continue
                    self.table_students.student_activities[self.table_students.student_id[student_name]].discard(activity_name)

            with open(get_activity_list_path(activity_name), "w", encoding="utf-8") as participants_list:
                participants_list.writelines(new_participants)
            
            if (len(activity_description)):
                self.activity_description_page_widgets[activity_name][0].setText(activity_description)
                with open(get_activity_description_path(activity_name), "w", encoding="utf-8") as description:
                    description.write(activity_description)

            self.input_win.close()
            self.open_activity_inner_page(activity_name)

    def open_delete_activity_window(self, activity_name):
        self.hide()

        self.input_win = InputWindow(self)
        self.input_win.setGeometry(400, 300, 400, 300)

        label = QLabel(f"Вы уверены, что хотите удалить мероприятие '{activity_name}'?", self)
        label.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label, alignment=Qt.AlignmentFlag.AlignCenter)

        button_go_back = QPushButton("Отмена")
        button_go_back.clicked.connect(lambda state, x = activity_name : self.return_from_activity_edit_or_deletion(x))
        self.input_win.layout.addWidget(button_go_back)

        button_commit = QPushButton("Подтвердить")
        button_commit.clicked.connect(lambda state, x = activity_name : self.delete_activity(x))
        self.input_win.layout.addWidget(button_commit)

        self.input_win.show()

    def delete_activity(self, activity_name):
        for i in range(len(self.activity_selection_buttons)):
            if (self.activity_selection_buttons[i].text() == activity_name):
                self.activity_selection_buttons.remove(self.activity_selection_buttons[i])
                break

        with open(get_activity_list_path(activity_name), "r", encoding="utf-8") as file:
            for student_name in file:
                student_name = delete_end_of_string(student_name)
                self.table_students.student_activities[self.table_students.student_id[student_name]].discard(activity_name)

        self.delete_activity_files(activity_name)
        self.show()
        self.open_activity_selection_page()

    def delete_activity_files(self, activity_name):
        os.remove(get_activity_list_path(activity_name))
        os.remove(get_activity_description_path(activity_name))

        new_activities = []
        with open(get_all_activities_path(), "r", encoding="utf-8") as all_activities:
            for line in all_activities:
                line = delete_end_of_string(line)
                if (line != activity_name):
                    new_activities.append(line + '\n')

        with open(get_all_activities_path(), "w", encoding="utf-8") as all_activities:
            all_activities.writelines(new_activities)

    def return_from_activity_edit_or_deletion(self, activity_name):
        if self.error_win:
            self.error_win.close()
        self.input_win.close()
        self.show()
        self.open_activity_inner_page(activity_name)
    
    def export_table(self):
        for i in range(self.table_students.table_model.rowCount()):
            self.table_students.set_value_in_cell(i, 3, len(self.table_students.student_activities[i]))
            self.table_students.set_value_in_cell(i, 4, set_to_str(self.table_students.student_activities[i]))
        
        connection = sqlite3.connect(get_db_path())
        wb = Workbook()
        ws = wb.active

        query = "SELECT * FROM students WHERE Количество_мероприятий > 0"
        cursor = connection.cursor()
        cursor.execute(query)
        data = cursor.fetchall()

        ws.cell(row=1, column=1, value="ФИО")
        ws.cell(row=1, column=2, value="Класс")
        ws.cell(row=1, column=3, value="Количество мероприятий")
        ws.cell(row=1, column=4, value="Мероприятия")
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_data in enumerate(row_data, start=1):
                if (col_index > 1):
                    ws.cell(row=row_index + 1, column=col_index - 1, value=cell_data)

        wb.save('Сводная таблица.xlsx')
        connection.close()

    def open_add_student_page(self):
        self.hide()

        self.input_win = InputWindow(self)
        self.input_win.setGeometry(400, 250, 450, 200)

        label_name = QLabel("Введите ФИО ученика:", self)
        label_name.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label_name, alignment=Qt.AlignmentFlag.AlignCenter)

        self.input_win.input_name = QLineEdit()
        self.input_win.input_name.setMinimumWidth(300)
        self.input_win.layout.addWidget(self.input_win.input_name, alignment=Qt.AlignmentFlag.AlignCenter)

        label_name2 = QLabel("Введите класс ученика:", self)
        label_name2.setFont(QFont("Helvetica [Cronyx]", 12))
        self.input_win.layout.addWidget(label_name2, alignment=Qt.AlignmentFlag.AlignCenter)

        self.input_win.input_name2 = QLineEdit()
        self.input_win.input_name2.setMinimumWidth(300)
        self.input_win.layout.addWidget(self.input_win.input_name2, alignment=Qt.AlignmentFlag.AlignCenter)
    
        button_commit = QPushButton("Подтвердить")
        button_commit.clicked.connect(self.add_student)
        self.input_win.layout.addWidget(button_commit)

        self.input_win.show()
    
    def add_student(self):
        student_name = self.input_win.input_name.text()
        class_name = self.input_win.input_name2.text()

        if ((student_name + " " + class_name) in self.table_students.student_id):
            self.error_win = ErrorWindow("Такой ученик уже существует")
            return

        ok = False
        with open(get_all_classes_path(), "r", encoding="utf-8") as all_classes:
            for cur_class_name in all_classes:
                cur_class_name = delete_end_of_string(cur_class_name)
                if (cur_class_name == class_name):
                    ok = True
        if (not ok):
            self.error_win = ErrorWindow(f"Класса '{class_name}' не существует.\nУбедитесь, что данные введены корректно.")
            return
        
        id = self.table_students.table_model.rowCount()
        self.table_students.table_model.insertRows(id, 1)
        self.table_students.set_value_in_cell(id, 1, student_name)
        self.table_students.set_value_in_cell(id, 2, class_name)
        self.table_students.student_id[student_name + " " + class_name] = id
        self.table_students.student_activities.append(set())

        with open(get_new_students_path(), "a", encoding="utf-8") as file:
            file.write(student_name + " " + class_name + "\n")

        self.input_win.close()