import os
import sqlite3
import openpyxl


def delete_end_of_string(string):
    if string[len(string) - 1] == '\n':
        string = string[0 : len(string) - 1]
    return string

def get_cwd():
    return os.getcwd()
    # if run from /src, add [0:-4] to the line above

def get_class_list_path(class_number):
    return get_cwd() + "\\data\\student_lists\\" + class_number + " список.xlsx"

def get_activity_list_path(activity_name):
    return get_cwd() + "\\data\\activity_lists\\" + activity_name + ".txt"

def get_activity_description_path(activity_name):
    return get_cwd() + "\\data\\activity_description\\" + activity_name + ".txt"

def get_all_classes_path():
    return get_cwd() + "\\data\\all_classes.txt"

def get_all_activities_path():
    return get_cwd() + "\\data\\all_activities.txt"

def get_db_path():
    return get_cwd() + "\\data\\student_database.db"

def get_new_students_path():
    return get_cwd() + "\\data\\student_lists\\new_students.txt"

def check_student_name(student_name):
    return (type(student_name) == str) and (student_name != "ФИО учащегося")

def create_student_database():
    connection = sqlite3.connect(get_db_path())
    cursor = connection.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS students
        (ID INTEGER, ФИО TEXT, Класс TEXT, Количество_мероприятий INTEGER, Мероприятия TEXT)
    """)

    with open(get_all_classes_path(), "r", encoding="utf-8") as all_classes:
        id = 0
        for cur_class_name in all_classes:
            cur_class_name = delete_end_of_string(cur_class_name)
            wb_obj = openpyxl.load_workbook(get_class_list_path(cur_class_name))
            sheet_obj = wb_obj.active
            for i in range(1, sheet_obj.max_row + 1):
                cell_obj = sheet_obj.cell(row = i, column = 3)
                student_name = cell_obj.value
                if (check_student_name(student_name)):
                    cursor.execute("INSERT INTO students VALUES ((?), (?), (?), 0, '')", (id, student_name, cur_class_name, ))
                    id += 1

    connection.commit()
    connection.close()
    
def set_to_str(st):
    if (len(st) == 0):
        return None
    s = str(st)
    return s[1:-1]