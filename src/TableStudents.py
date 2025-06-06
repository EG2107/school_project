from PyQt6.QtWidgets import QTableView
from PyQt6.QtSql import QSqlDatabase, QSqlTableModel
from functions import delete_end_of_string, get_class_list_path, get_activity_list_path, get_all_classes_path, get_all_activities_path, get_db_path, get_new_students_path, check_student_name
import openpyxl
from os.path import exists

class TableStudents(QTableView):
    def __init__(self):
        super().__init__()

        self.db = QSqlDatabase.addDatabase("QSQLITE")
        self.db.setDatabaseName(get_db_path())
        self.db.open()

        self.table_model = QSqlTableModel(None, self.db)
        self.table_model.setTable("students")
        self.table_model.select()
        while (self.table_model.canFetchMore()):
            self.table_model.fetchMore()

        self.init_additional_data()

        self.setModel(self.table_model)
        self.move(100, 100)
        self.resize(800, 600)
        self.setColumnWidth(0, 15)
        self.setColumnWidth(1, 300)
        self.setColumnWidth(2, 70)
        self.setColumnWidth(3, 160)
        self.setColumnWidth(4, 800)
        self.hide()

    def init_additional_data(self):
        self.student_activities = []
        self.student_id = dict()

        id = 0
        with open(get_all_classes_path(), "r", encoding="utf-8") as all_classes:
            for cur_class_name in all_classes:
                cur_class_name = delete_end_of_string(cur_class_name)
                wb_obj = openpyxl.load_workbook(get_class_list_path(cur_class_name))
                sheet_obj = wb_obj.active
                for i in range(1, sheet_obj.max_row + 1):
                    cell_obj = sheet_obj.cell(row = i, column = 3)
                    student_name = cell_obj.value
                    if (check_student_name(student_name)):
                        self.student_id[student_name + " " + cur_class_name] = id
                        self.student_activities.append(set())
                        id += 1

        if not exists(get_new_students_path()):
            with open(get_new_students_path(), "w", encoding="utf-8") as new_students:
                pass
        with open(get_new_students_path(), "r", encoding="utf-8") as new_students:
            for line in new_students:
                self.student_id[delete_end_of_string(line)] = id
                self.student_activities.append(set())
                id += 1

        with open(get_all_activities_path(), "r", encoding="utf-8") as all_activities:
            for cur_activity_name in all_activities:
                cur_activity_name = delete_end_of_string(cur_activity_name)
                with open(get_activity_list_path(cur_activity_name), "r", encoding="utf-8") as file:
                    for student_name in file:
                        student_name = delete_end_of_string(student_name)
                        self.student_activities[self.student_id[student_name]].add(cur_activity_name)

    def get_value_in_cell(self, row, column):
        index = self.table_model.index(row, column)
        return self.table_model.data(index)

    def add_value_in_cell(self, row, column, value):
        index = self.table_model.index(row, column)
        old_value = self.table_model.data(index)
        self.table_model.setData(index, old_value + value)
        self.table_model.submit()
    
    def set_value_in_cell(self, row, column, value):
        index = self.table_model.index(row, column)
        self.table_model.setData(index, value)
        self.table_model.submit()